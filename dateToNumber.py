import dearpygui.dearpygui as dpg
import openpyxl
import tkinter
from tkinter import filedialog

FILENAME = ""
DATE_COLMUN = 0
NUMBER_COLMUN = 0

dpg.create_context() #dearpyguiを使う宣言。必須。
dpg.create_viewport() #viewportという描画スペースの宣言。必須。
dpg.setup_dearpygui() #これも必須。


def choose_file(sendar,data):
    global FILENAME
    root = tkinter.Tk()
    root.withdraw()  # ルートウィンドウを非表示に設定
    filename = filedialog.askopenfilename()
    root.destroy()  # 非表示になっているtkinterのウィンドウを削除する
    FILENAME = filename
    print(FILENAME)
    
def date_to_number():
    global FILENAME
    global DATE_COLMUN
    global NUMBER_COLMUN
    
    wb = openpyxl.load_workbook(FILENAME)
    ws = wb["Sheet1"]
    MAX_ROW = ws.max_row
    for i in range(2,MAX_ROW+1):
        date = ws.cell(i,int(DATE_COLMUN))
        date = str(date.value).replace('-','')
        date = date[0:8]
        ws.cell(i,int(NUMBER_COLMUN)).value = int(date)
        print(date)
    wb.save(FILENAME)
    wb.close()
    
def set_date_colmun(sendar,data):
    global DATE_COLMUN
    DATE_COLMUN = data
    print(DATE_COLMUN)
    
def set_number_colmun(sendar,data):
    global NUMBER_COLMUN
    NUMBER_COLMUN = data
    print(NUMBER_COLMUN)

with dpg.window(width=400,height=400):#描画スペースの中にwindowを作る。
    dpg.add_text("Choose Excel File")#上記windowの中にテキストを表示。
    dpg.add_button(label="push",callback=choose_file)
    dpg.add_input_text(label="input date colmun",callback=set_date_colmun)
    dpg.add_input_text(label="input number colmun",callback=set_number_colmun)
    dpg.add_button(label="exe",callback=date_to_number)

dpg.show_viewport()#viewportを表示。必須。
dpg.start_dearpygui()#DearPyGuiのスタート。必須
dpg.destroy_context()#DearPyGuiのクリンナップ。必須
